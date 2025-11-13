import openpyxl
import json

# Carica il file Excel (data_only=True per leggere valori invece di formule)
wb = openpyxl.load_workbook('DOCUMENTAZIONE/Lista Modelli.xlsx', data_only=True)

result = {
    'schede': wb.sheetnames,
    'liste': {},
    'modelli': {}
}

# Analizza scheda Liste
liste = wb['Liste']
result['liste']['tipologie'] = []
result['liste']['marche'] = []

for row in liste.iter_rows(min_row=1, max_row=liste.max_row, min_col=1, max_col=2, values_only=True):
    if row[0] and str(row[0]).strip():
        result['liste']['tipologie'].append(str(row[0]).strip())
    if row[1] and str(row[1]).strip():
        result['liste']['marche'].append(str(row[1]).strip())

# Analizza scheda Modelli
modelli = wb['Modelli']
header = [str(cell.value) if cell.value else '' for cell in modelli[1]]
result['modelli']['header'] = header
result['modelli']['righe'] = []

for row in modelli.iter_rows(min_row=2, max_row=modelli.max_row, values_only=True):
    if any(row):  # Salta righe vuote
        # Converti tutti i valori in stringhe per evitare problemi JSON
        cleaned_row = [str(val) if val is not None else '' for val in row]
        result['modelli']['righe'].append(cleaned_row)

# Salva output
print(json.dumps(result, indent=2, ensure_ascii=False))
